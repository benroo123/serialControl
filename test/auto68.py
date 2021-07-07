import serial
import time
import xlwt
from openpyxl import load_workbook, Workbook
import importlib
import os
from loguru import logger as log
import sys


"""
the method is used to check the at commands which should have OK as the reply
machine is a serial instance
cmd is the command corresponding to the output
origin is the time the command is sent
"""
def see_ok(machine, cmd, origin):
    diff = origin - time.time()
    # use one second to get the reply
    while diff <= 1:
        diff = time.time() - origin
        try:
            mode = machine.readline().decode()[:-2]
            if mode == "OK":
                print(cmd[:-2], ":", mode, "  %.4f" % diff)
                return "OK", diff
        except Exception:
            print(cmd, "Couldn't get the output.")
            return [-1], diff
    return "nothing here", diff


"""

"""
def see_request(machine, cmd, origin):
    diff = origin - time.time()
    reply = []  # the list to take the reply

    try:
        feedback = machine.readline().decode()[:-2]
        # use one second to take the reply
        # if the reply is not empty, put it into the list
        # if read OK then quit the while loop directly
        while diff <= 1:
            diff = time.time() - origin
            if feedback != " ":
                reply.append(feedback)
            if reply[-1] == "OK":
                break
            if reply[-1] == config.error2:
                break
            feedback = machine.readline().decode()[:-2]
        print(cmd[:-2], ":")
        for e in reply:
            print(e)
        print("%.4f" % diff)
        if reply:
            return reply, diff
        else:
            return "nothing here", diff
        # return reply, diff
    except Exception:
        print(cmd, "Couldn't get the output.")
        return [-1], diff


def send(machine, formatter, check_list, path_excel):
    i = 1  # the int used to walk through the list
    row = 1
    while i <= len(check_list):
        # wait for an extra second after sending the reset command
        # there are too many replies by sending this command
        if check_list[i - 2] == "AT+DEF\r\n":
            time.sleep(1)
        # send the command
        machine.write(check_list[i - 1].encode("utf-8"))
        o = time.time()
        formatter[1].write(row, 0, check_list[i - 1])  # write the command into the form
        machine.reset_input_buffer()
        if i % 2 != 0:
            # write the reply and time into different slots
            result_ok = see_ok(machine, check_list[i - 1], o)
            formatter[1].write(row, 1, result_ok[0])
            formatter[1].write(row, 2, result_ok[1])
            row += 1
        else:
            result2 = see_request(machine, check_list[i - 1], o)
            formatter[1].write(row, 2, result2[1])
            for e in result2[0]:
                formatter[1].write(row, 1, e)
                row += 1

            print("=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=")
        i += 1
        time.sleep(2)

    formatter[0].save(path_excel)


"""
The method to make connection with the temp_config.py
"""
def import_config(path_config):
    global config
    log.info("load config {}".format(path_config))
    module_name_path = "temp_config.py"
    os.system("cp {0}  {1}".format(path_config, module_name_path))

    config = importlib.import_module(module_name_path[:-3])


def write_to_form(key_list):
    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = 'SimSun'
    wb = Workbook()
    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('my doc')

    i = -1
    for word in key_list:
        i += 1
        worksheet.write(0, i, word.upper(), style)

    return workbook, worksheet


def initiate(port, baudrate):
    test_machine = serial.Serial(port, baudrate)
    return test_machine


if __name__ == "__main__":
    # print(sys.argv[0][:-9])
    import_config(sys.argv[0])
    send(initiate(config.port, config.baudrate), write_to_form(config.key_list), config.command_seq, config.path_excel)
