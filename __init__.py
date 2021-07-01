import serial
import time
import xlwt
from openpyxl import load_workbook, Workbook
import importlib
import os
from loguru import logger as log
import sys

# set the parameters of the serial
port = "COM5"
baudrate = 115200


# commands will be used
exiT = "AT+EXIT\r\n"
tri_plus = "+++\r\n"
reset = "AT+DEF\r\n"
request1 = "AT+SYMBOL=?\r\n"
request2 = "AT+FREQ?\r\n"

# the list of all the commands will be called in this test
command_seq = [exiT, request1, reset, request2, tri_plus, request1]

# the head of the excel form
key_list = ["指令", "返回", "时间"]

# the path to store the result
path_excel = "./result_{}.xls".format(int(time.time()))


"""
the method is used to check the at commands which should have OK as the reply
machine is a serial instance
cmd is the command corresponding to the output
origin is the time the command is sent
"""
def see_ok(machine, cmd, origin):
    diff = origin - time.time()
    # use one second to get the reply
    while diff <= 1:
        diff = time.time() - origin
        try:
            mode = machine.readline().decode()[:-2]
            if mode == "OK":
                print(cmd[:-2], ":", mode, "  %.4f" % diff)
                return "OK", diff
            else:
                return "no output", diff
        except Exception:
            print(cmd, "Couldn't get the output.")
            return [-1], diff


"""

"""
def see_request(machine, cmd, origin):
    diff = origin - time.time()
    reply = []  # the list to take the reply

    try:
        feedback = machine.readline().decode()[:-2]
        # use one second to take the reply
        # if the reply is not empty, put it into the list
        # if read OK then quit the while loop directly
        while diff <= 1:
            diff = time.time() - origin
            if feedback != " ":
                reply.append(feedback)
            if reply[-1] == "OK":
                break
            feedback = machine.readline().decode()[:-2]
        print(cmd[:-2], ":")
        for e in reply:
            print(e,)
        print("%.4f" % diff)
        if reply:
            return reply, diff
        else:
            return "nothing here", diff
        # return reply, diff
    except Exception:
        print(cmd, "Couldn't get the output.")
        return [-1], diff


def send(machine, formatter, check_list, path_excel):
    i = 1  # the int used to walk through the list
    row = 1
    while i <= len(check_list):
        # wait for an extra second after sending the reset command
        # there are too many replies by sending this command
        if check_list[i - 2] == "AT+DEF\r\n":
            time.sleep(1)
        # send the command
        machine.write(check_list[i - 1].encode("utf-8"))
        o = time.time()
        formatter.write(row, 0, check_list[i - 1]) # write the command into the form
        if i % 2 != 0:
            # write the reply and time into different slots
            result_ok = see_ok(machine, check_list[i - 1], o)
            formatter.write(row, 1, result_ok[0])
            formatter.write(row, 2, result_ok[1])
            row += 1
        else:
            machine.reset_input_buffer()
            result2 = see_request(machine, check_list[i - 1], o)
            formatter.write(row, 2, result2[1])
            for e in result2[0]:
                formatter.write(row, 1, e)
                row += 1

            print("=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=")
        i += 1
        time.sleep(2)

    formatter.save(path_excel)


"""
The method to make connection with the temp_config.py
"""
def import_config(path_config):
    global config
    log.info("load config {}".format(path_config))
    module_name_path = "temp_config.py"
    os.system("cp {0}  {1}".format(path_config, module_name_path))

    config = importlib.import_module(module_name_path[:-3])


def write_to_form(key_list):
    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = 'SimSun'
    wb = Workbook()
    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('my doc')

    i = -1
    for word in key_list:
        i += 1
        worksheet.write(0, i, word.upper(), style)

    return worksheet


def initiate(port, baudrate):
    testMachine = serial.Serial(port, baudrate)
    return testMachine


send(initiate(port, baudrate), write_to_form(key_list), command_seq, path_excel)